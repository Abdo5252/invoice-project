import pandas as pd
import io
import openpyxl
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def create_output_excel(processed_invoices, template_file=None):
    """
    Create a new Excel file with two sheets:
    1. Headers: Contains invoice header information
    2. Items: Contains product details from all invoices
    
    Args:
        processed_invoices: List of dictionaries containing processed invoice data
        template_file: The uploaded template Excel file object (optional, not used in the new approach)
        
    Returns:
        BytesIO object containing the Excel file
    """
    # Create a new workbook for output
    output = io.BytesIO()
    
    # If there are no invoices, return an empty file with default sheets
    if not processed_invoices:
        workbook = openpyxl.Workbook()
        # Create Header sheet
        header_sheet = workbook.active
        header_sheet.title = "Header"
        header_sheet.append(["Document Type", "Document Number", "Document Date", "Customer Code", 
                            "Currency Code", "Exchange Rate", "Extra Discount", "Activity Code"])
        
        # Create Items sheet
        items_sheet = workbook.create_sheet(title="Items")
        items_sheet.append(["Document Number", "Description", "Unit Type", "Quantity", 
                           "Unit Price", "Discount Amount", "Value Difference", "Item Discount"])
        
        workbook.save(output)
        output.seek(0)
        return output
    
    # Create a new workbook
    workbook = openpyxl.Workbook()
    
    # Set up Headers sheet
    header_sheet = workbook.active
    header_sheet.title = "Header"
    header_sheet.append(["Document Type", "Document Number", "Document Date", "Customer Code", 
                        "Currency Code", "Exchange Rate", "Extra Discount", "Activity Code", "Total Amount"])
    
    # Set up Items sheet
    items_sheet = workbook.create_sheet(title="Items")
    items_sheet.append(["Document Number", "Internal Code", "Description", "Unit Type", "Quantity", 
                       "Unit Price", "Discount Amount", "Value Difference", "Item Discount"])
    
    # Always use current date for document date
    current_date = datetime.now().strftime("%m/%d/%Y")
    
    # Process each invoice
    for invoice in processed_invoices:
        # Add header row for this invoice
        # Set exchange rate based on currency: 52 for USD, 0 for others
        currency = invoice.get('currency', '')
        exchange_rate = "0"  # Default exchange rate
        
        if currency == 'USD':
            exchange_rate = "52"
        
        header_sheet.append([
            "I",  # Document Type (Always "I" as per requirements)
            invoice.get('invoice_number', ''),  # Document Number
            current_date,  # Document Date (always use current date)
            invoice.get('customer_code', ''),  # Customer Code
            invoice.get('currency', ''),  # Currency Code
            exchange_rate,  # Exchange Rate (based on currency)
            "0",  # Extra Discount (default to 0 as per requirements)
            "",   # Activity Code (empty if not available as per requirements)
            invoice.get('total_amount', 0)  # Total Invoice Amount
        ])
        
        # Add product rows for this invoice
        if 'products' in invoice and invoice['products']:
            for product in invoice['products']:
                # Use document_number from product if available (from the structured table extraction)
                # Otherwise use the invoice's invoice_number
                doc_number = product.get('document_number', invoice.get('invoice_number', ''))
                
                items_sheet.append([
                    doc_number,  # Document Number (links to header)
                    "1",  # Internal Code (always 1 as requested)
                    product.get('description', ''),  # Description
                    product.get('unit_type', ''),  # Unit Type (leave blank if not found as per requirements)
                    product.get('quantity', ''),  # Quantity
                    product.get('unit_price', ''),  # Unit Price
                    "0",  # Discount Amount (default to 0 as per requirements)
                    "0",  # Value Difference (default to 0 as per requirements)
                    "0"   # Item Discount (default to 0 as per requirements)
                ])
    
    # Auto-size columns in both sheets
    for sheet in [header_sheet, items_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Find the maximum content length in each column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (with some padding)
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)
    
    return output

def extract_invoice_date(df):
    """
    Extract invoice date from the dataframe.
    
    Args:
        df: DataFrame with string values
        
    Returns:
        Extracted date as string in YYYY-MM-DD format or None if not found
    """
    # Keywords that might precede an invoice date
    date_keywords = [
        'date', 'invoice date', 'issued on', 'تاريخ', 'تاريخ الفاتورة'
    ]
    
    # Common date formats (DD/MM/YYYY, MM/DD/YYYY, YYYY-MM-DD, etc.)
    date_patterns = [
        r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',  # DD/MM/YYYY or MM/DD/YYYY
        r'(\d{2,4}[/-]\d{1,2}[/-]\d{1,2})',  # YYYY-MM-DD
    ]
    
    # Search for each keyword in the dataframe
    for keyword in date_keywords:
        for i in range(len(df)):
            for j in range(len(df.columns)):
                cell = df.iloc[i, j].lower()
                if keyword.lower() in cell:
                    # Check this cell for date patterns
                    for pattern in date_patterns:
                        match = re.search(pattern, cell)
                        if match:
                            # Try to parse the date
                            try:
                                # This is simplified - in production would need more robust date parsing
                                return match.group(1)
                            except:
                                pass
                    
                    # Check the cell to the right
                    if j + 1 < len(df.columns):
                        right_cell = df.iloc[i, j + 1]
                        for pattern in date_patterns:
                            match = re.search(pattern, right_cell)
                            if match:
                                try:
                                    return match.group(1)
                                except:
                                    pass
    
    # If date not found, return None
    return None
